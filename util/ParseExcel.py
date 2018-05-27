# -*- coding = utf-8 -*-
import openpyxl
from openpyxl.styles import Border, Side, Font
import time


class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def load_workbook(self, excel_path_and_name):
        # 将Excel文件加载到内存，并获取workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excel_path_and_name)
        except Exception as e:
            raise e
        self.excelFile = excel_path_and_name
        return self.workbook

    def get_sheetbyname(self, sheetname):
        # 根据sheet名获取该sheet对象
        try:
            sheet = self.workbook[sheetname]
            return sheet
        except Exception as e:
            raise e

    def get_sheetbyindex(self, sheetindex):
        # 根据sheet的索引号获取该sheet对象
        try:
            sheetname = self.workbook.sheetnames[sheetindex]
        except Exception as e:
            raise e
        sheet = self.workbook[sheetname]
        return sheet

    def get_end_row(self, sheet):
        # 获取sheet中有数据区域的结束行号
        return sheet.max_row

    def get_end_col(self, sheet):
        # 获取sheet中有数据区域的结束列号
        return sheet.max_column

    def get_start_row(self, sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_row

    def get_start_col(self, sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_column

    def get_row(self, sheet, row_number):
        # 获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple,下标从1开始，sheet.row[1]表示第一列
        try:
            list_row = []
            for i in list(sheet.rows)[row_number-1]:
                list_row.append(i.value)
            return list_row
        except Exception as e:
            raise e

    def get_column(self, sheet, column_number):
        # 获取sheet表中某一列
        try:
            list_column = []
            for i in list(sheet.columns)[column_number-1]:
                list_column.append(i.value)
            return list_column
        except Exception as e:
            raise e

    def get_cell_value(self, sheet, coordinate=None, row_number=None, column_number=None):
        # 根据单元格所在的位置索引获取单元格的值，下标从1开始，sheet.cell(row = 1,column = 1).value,表示第一行第一列的值
        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and row_number is not None and column_number is not None:
            try:
                return sheet.cell(row=row_number, column=column_number).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def get_cell_object(selfself, sheet, coordinate=None, row_number=None, column_number=None):
        # 获取某个单元格的对象，可以根据单元格所在位置的数字索引，也可以直接根据Excel中单元格的编码及坐标
        # 如getCellobject(sheet,coordinate='AI')or getcellobject(sheet,row_number=1,column_numbe=1)
        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate is None and row_number is not None and column_number is not None:
            try:
                return sheet.cell(row=row_number, column=column_number)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def write_cell(self, sheet, content, coordinate=None, row_number=None, column_number=None, style=None):
        # 根据单元格在excel中彪马坐标或者数字索引坐标向单元格中写入数据，下标从1开始，参数style表示字体颜色
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and row_number is not None and column_number is not None:
            try:
                sheet.cell(row=row_number, column=column_number).value = content
                if style:
                    sheet.cell(row=row_number, column=column_number).font = Font(color=self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def write_cell_currenttime(self, sheet, coordinate=None, row_number=None, column_number=None):
        # 写入当前的时间，下标从1开始
        now = int(time.time())     # 显示为时间戳
        time_array = time.localtime(now)
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time_array)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = current_time
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and row_number is not None and column_number is not None:
            try:
                sheet.cell(row=row_number, column=column_number).value = current_time
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell")


if __name__ == '__main__':
    pe = ParseExcel()
    pe.load_workbook(u'F:\\DataDriverFrameWork\\testData\\联系人.xlsx')
    print(u'通过名称获取sheet对象的名字', pe.get_sheetbyname(u'add1').title)
    print(u'通过index序列号获取sheet对象的名字', pe.get_sheetbyindex(0).title)
    sheet = pe.get_sheetbyindex(0)
    # print(type(sheet))
    # print(pe.get_end_row(sheet))
    # print(pe.get_end_col(sheet))
    print(pe.get_row(sheet, 1))
    print(pe.get_column(sheet, 2))
    print(pe.get_cell_value(sheet, row_number=1, column_number=1))
    # pe.write_cell(sheet=sheet, content=u'我爱祖国', row_number=2, column_number=10)
    # pe.write_cell_currenttime(sheet=sheet, row_number=2, column_number=11)
